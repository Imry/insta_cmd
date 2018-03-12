#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, colors
from openpyxl.utils.cell import get_column_letter

from g import *


STYLE_BOLD = NamedStyle(name="bold")
STYLE_BOLD.font = Font(bold=True)
STYLE_WRAP = NamedStyle(name="wrap")
STYLE_WRAP.alignment = Alignment(wrapText=True)
STYLE_LINK = NamedStyle(name="Hyperlink")
STYLE_LINK.font = Font(color=colors.BLUE)
STYLE_LINK.alignment = Alignment(wrapText=True)


def link(url):
    return ['=HYPERLINK("%s", "%s")' % (url, url), STYLE_LINK]


def create_write(sheet, col, row):
    def writer(c, r, data, style=None):
        c = sheet.cell(column=1+row + r, row=1+col + c, value=data)
        if style:
            c.style = style
    return writer


def write_info(sheet, data, col, row):
    writer = create_write(sheet, col, row)

    writer(0, 0, 'IULogin', STYLE_BOLD)
    writer(0, 1, 'IUFoto', STYLE_BOLD)
    writer(0, 2, 'IUName', STYLE_BOLD)
    writer(0, 3, 'IUSite', STYLE_BOLD)
    writer(0, 4, 'IUNote', STYLE_BOLD)

    writer(1, 0, data.username)
    writer(2, 0, *link(data.username))

    writer(1, 1, *link(data.img))

    writer(1, 2, data.name, STYLE_WRAP)

    writer(1, 3, *link(data.url))

    writer(1, 4, data.bio)


def chunks(l):
    """Yield successive n-sized chunks from l."""
    for i in range(0, len(l), EXCEL_MAX_LINES):
        yield l[i:i + EXCEL_MAX_LINES]


def write_user_list(sheet, data, col, row, title, key):
    shift_col = 0
    writer = create_write(sheet, col, row)
    writer(0, 0 + shift_col * 3, title, STYLE_BOLD)

    l = getattr(data, key, None)
    if not l:
        return shift_col

    writer(1, 0, l.count)
    writer(2, 0, 'IULogin', STYLE_BOLD)
    writer(2, 1, 'IUName', STYLE_BOLD)
    writer(2, 2, 'IUFoto', STYLE_BOLD)
    for shift_col, c in enumerate(chunks(l.data)):
        writer(2, 0 + shift_col * 4, 'IULogin', STYLE_BOLD)
        writer(2, 1 + shift_col * 4, 'IUName', STYLE_BOLD)
        writer(2, 2 + shift_col * 4, 'IUFoto', STYLE_BOLD)

        for idx, f in enumerate(c):
            writer(3 + idx, 0 + shift_col * 4, f.username)
            writer(3 + idx, 1 + shift_col * 4, f.name, STYLE_WRAP)
            writer(3 + idx, 2 + shift_col * 4, *link(f.img))
    return shift_col


def write_media(sheet, data, col, row):
    writer = create_write(sheet, col, row)

    writer(0, 0, 'UIPost', STYLE_BOLD)
    writer(0, 1, data.media.count)
    writer(1, 0, 'IPUrl', STYLE_BOLD)
    writer(1, 1, 'IPLocation', STYLE_BOLD)
    writer(1, 2, 'IPDate', STYLE_BOLD)
    writer(1, 3, 'IPLike', STYLE_BOLD)
    writer(1, 4, 'IPFoto', STYLE_BOLD)
    writer(1, 5, 'IPComments', STYLE_BOLD)

    shift_row = 2
    for p in data.media.data:
        writer(shift_row, 0, *link(p.url))
        location = p.location_str
        if location != '':
            location = location.split('\n')
            writer(shift_row, 1, location[0], STYLE_WRAP)
            writer(shift_row + 1, 1, *link(location[1]))
        writer(shift_row, 2, p.time_str)
        writer(shift_row, 3, p.likes)
        m = p.media_str.split('\n')
        for idx, mm in enumerate(m):
            writer(shift_row + idx, 4, *link(mm))
        writer(shift_row, 5, p.comment_str, STYLE_WRAP)

        shift_row += max([1, len(location), len(m)])


def save(d, fn):
    book = openpyxl.Workbook()
    sheet = book.create_sheet(d.username, 0)

    write_info(sheet, d, 0, 0)
    write_media(sheet, d, 0, 6)
    shift_col_following = write_user_list(sheet, d, 0, 13, 'IULoginIn',  'following')
    shift_col_followers = write_user_list(sheet, d, 0, 17 + shift_col_following * 4, 'IULoginOut', 'followers')

    for i in range(1, 17 + shift_col_following * 4 + (shift_col_followers+1) * 4):
        sheet.column_dimensions[get_column_letter(i)].width = 50
    for i in [6, 13]:
        sheet.column_dimensions[get_column_letter(i)].width = 10
    for i in range(13, 13 + (shift_col_following+1) * 4, 4):
        sheet.column_dimensions[get_column_letter(i)].width = 10
    for i in range(17 + shift_col_following * 4, 17 + shift_col_following * 4 + (shift_col_followers+1) * 4, 4):
        sheet.column_dimensions[get_column_letter(i)].width = 10

    book.save(fn)
